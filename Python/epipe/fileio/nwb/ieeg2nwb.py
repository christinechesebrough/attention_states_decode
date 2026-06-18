#!/usr/bin/env python

# ieeg2nwb.py
# A python-based script meant to convert edf, TDT and ecog.mat formatted data into an
# NWB file
#
# Noah Markowitz
# Human Brain Mapping Lab
# North Shore University Hospital
# April 2021

# Imports
import numpy as np
import pandas as pd
import os
import re
import tdt
import uuid
import gzip
from itertools import compress
from mne.io import read_raw_edf
from datetime import datetime
from dateutil.tz import tzlocal
from hdmf.backends.hdf5.h5_utils import H5DataIO
from hdmf.common.table import DynamicTable
import json
import yaml
import argparse
import sys
from openpyxl import load_workbook
from .utils import load_nwb_settings
from .messages import example_usage, additional_notes
from ..tdt import getTDTStore, read_tdt_ttls
from sklearn.mixture import GaussianMixture
from colorama import Back, Style

from ...utilities import recurse_np2list

try:
    from mne.externals.pymatreader import read_mat
except:
    from pymatreader import read_mat

# TODO
#   - "make BIDS" option to create json sidecars for BIDS
#   - "create_path" to create output directory if it does not exist
#   - All above should work in
#       * Command-line
#       * params file
#       * gui
#       * batch
#   - If correspondence sheet as blank cells that contain only whitespace, remove the whitespaces

class ieeg2nwb:

    def __init__(self,session_description = 'None',session_id = 'None'):

        params = load_nwb_settings()

        # Meta-data
        self.file_type = ''
        self.start_time = datetime.strptime(params['meta_data']['date'],"%Y-%m-%d %H:%M:%S").replace(tzinfo=tzlocal())  #datetime.now(tzlocal())
        for metd in params['meta_data'].keys():
            setattr(self,metd,params['meta_data'][metd])

        # NWB objects
        self.identifier = str(uuid.uuid4())
        self.device_types = params['devices']
        self.device = None
        self.subject = None
        self._default_subject_info = params['subject']
        self.events = None
        self.electrode_group = []
        self.acquisitions = []
        self.nwb = None
        self.nwbhdf5io = None
        self.extra_files = []
        self.keywords = []
        self.experimenter = []
        self.blockfile = None
        self.create_path = False

        # Electrodes table
        self.electrode_table = params['electrode_table']
        self.electrode_table['dynamic_columns'] = None
        self.electrode_table['table_regions'] = None
        self.electrode_table['table'] = None
        self.labelfile = {'orig': None, 'cleaned': None}
        self.electrode_table['xltek_chans'] = None
        self.electrode_table['tdt_chans'] = None
        self.ieeg_elecs = params['lfp_specs']

        # Timeseries data
        self.raw_data = {
            'orig': None,
            'neural': {'data': None, 'fs': None,'labels': None, 'units': 'unknown'},
            'analog': [],
            'digital': []
        }
        self.other_neural_ts = [] # Each element must be dict of form {'name': <name>, 'fs': <fs>, 'data': [], 'comments': "", 'description': "", 'units': ""}

        # Additional notes
        self.annotations = {'timestamps': [], 'notes': []}

        # Handling imports

        self.tdt_eeg_chans = params['raw']['tdt']['neuro_channels']
        self.annots2ignore = params['raw']['edf']['annotations_to_ignore']

    # Create subject object for nwb file
    def create_subject(self,subject_id=None,sex=None,species=None,age=None,subject_description=None):
        from pynwb.file import Subject
        self.subject = Subject(
            age=self._default_subject_info['age'] if age is None else 'P' + str(age) + 'Y',
            sex=self._default_subject_info['sex'] if sex is None else sex,
            species=self._default_subject_info['species'] if species is None else species,
            subject_id=self._default_subject_info['subject_id'] if subject_id is None else subject_id,
            description=self._default_subject_info['description'] if subject_description is None else subject_description
        )

        return self.subject

    # Create device object
    def create_device(self, name='Unknown', description=None, manufacturer=None):
        from pynwb.device import Device
        possible_devices = self.device_types
        # Check if name provided is in any of the possible devices set and others are empty
        if description == None and manufacturer == None and name.lower() in possible_devices.keys():
            my_dev_dict = possible_devices[name]
            if 'search' in my_dev_dict.keys(): my_dev_dict.pop('search')
            dev = Device(**my_dev_dict)
        else:
            dev = Device(name, description, manufacturer)

        self.device = dev
        return self.device

    # Read labelfile and clean up a bit
    def read_labelfile(self,labelfile):

        # If no labelfile provided and ecog file is used then use available info in file
        if labelfile == 'ecog':
            df = pd.DataFrame(self.raw_data['orig']['labelimport'])
            df = df.rename(columns={'Channeltype': 'spec'})
            df['bad'] = df['bad'].fillna(0)
            df['soz'] = df['soz'].fillna(0)
            df['spikey'] = df['spikey'].fillna(0)
            block_bad = self.raw_data['orig']['spike_chans'] + self.raw_data['orig']['bad_chans']
            df['artifact_block'] = df['labels'].isin(block_bad).astype('int')
        elif labelfile.endswith('.csv'):
            df = pd.read_csv(labelfile)
        else:
            df = pd.read_excel(labelfile, sheet_name=0,engine='openpyxl')

        self.labelfile['orig'] = df

    #region Read Extra Files
    # Read a text file that's not the labelfile
    def read_extra_file(self, extra_file):
        if extra_file.endswith('.csv'):
            return pd.read_csv(extra_file)
        elif extra_file.endswith('.xlsx'):
            return pd.read_excel(extra_file)
        elif extra_file.endswith('.mat'):
            
            mat = read_mat(extra_file)
            mat.pop('__header__')
            mat.pop('__version__')
            mat.pop('__globals__')

            mydict = recurse_np2list(mat)
                   
            js = json.dumps(mydict)
            mydict = json.loads(js)
            for k in mydict.keys():
                mydict[k] = json.dumps(mydict[k])

            return pd.DataFrame(mydict, index=[0])

        else:
            print('unsupported file type!')

    # Convert a pandas DataFrame to DynamicTable in ScratchData
    def add_scratch(self,df,name,description='a raw file from experiment'):

        # Make sure df is correct
        dyn_columns = []
        cols2drop = []
        for colname, contents in df.iteritems():
            if colname.startswith('Unnamed'):
                cols2drop.append(colname)
            else:
                dyn_columns.append({'name': colname, 'description': 'auto-generated'})

        df = df.drop(columns=cols2drop)

        # Make into a DynamicTable
        dyn = DynamicTable(name=name, description=description).from_dataframe(
            df,
            name=name,
            table_description=description,
            columns=dyn_columns
        )

        # Make the ScratchData object

        self.scratch.append(dyn)

    def add_file(self, df, name, description='a raw file from experiment'):

        # Make sure df is correct
        dyn_columns = []
        cols2drop = []
        for colname, contents in df.iteritems():
            if colname.startswith('Unnamed'):
                cols2drop.append(colname)
            else:
                dyn_columns.append({'name': colname, 'description': 'auto-generated'})

        df = df.drop(columns=cols2drop)

        # Make into a DynamicTable
        dyn = DynamicTable(name=name, description=description).from_dataframe(
            df,
            name=name,
            table_description=description,
            columns=dyn_columns
        )

        self.extra_files.append(dyn)

    #endregion

    # Validate channels of edf and tdt recordings
    # remove those without a label, i.e. []
    def validate_chans(self, chansDf, chanCol):
        dfcols = list(chansDf.columns)
        r = re.compile(chanCol, re.IGNORECASE)
        chanNumCol = list(filter(r.match, dfcols))[0]
        validation = []
        for idx, r in chansDf.iterrows():
            chan_check = np.logical_and(r[chanNumCol] != '[]', r[chanNumCol] == r[chanNumCol])
            validation.append(chan_check)
                
        if np.sum(np.invert(validation)) != 0:
            print(Back.YELLOW + '{:d} invalid channels found in the label file!'.format(np.sum(np.invert(validation))))
            print(Style.RESET_ALL)

        valid_chans = chansDf.loc[validation, [dfcols[0], chanNumCol]]
        valid_chans.columns = ['label', 'chan']
        return valid_chans, validation

    # Format labelfile before converting it to DynamicTable
    def clean_labelfile(self,file_type=None,labels=None):
        df = self.labelfile['orig']
        neural_data = self.raw_data['neural']['data']
        
        # Warn if length of the neural data and labels is not the same
        # Will be corrected below, but should be looked into
        if neural_data.shape[0] != len(df):
            print(Back.YELLOW + 'Number of channels in neural data and channel label file don''t match!')
            print(Style.RESET_ALL)

        if file_type == None:
            file_type = self.file_type

        spec_column_settings = self.electrode_table.get('columns').get('spec')

        # First check for the label column
        # A few steps only apply to edf and TDT files
        if file_type != 'ecog':
            # Start by getting rid of any rows without labels
            nullRows = df.loc[:,df.columns[0]].isnull()
            df = df.loc[~nullRows,:]

            # Get rid of all channels without "spec" value
            r = re.compile(spec_column_settings.get('search'),re.IGNORECASE)
            specColName = list(filter(r.match, list(df.columns)))[0]
            rows2delete = []
            for idx, row in df.iterrows():
                if not isinstance(row[specColName],str) or row[specColName] == '[]':
                    rows2delete.append(idx)

            df = df.drop(index=rows2delete)
            
            if len(rows2delete) != 0:
                print(Back.YELLOW + 'Removed {:d} channels with missing ''spec'' field!'.format(len(rows2delete)))
                print(Style.RESET_ALL)

            # Remove more rows depending on neural data file type
            if file_type == 'edf':
                valid_chans, validation = self.validate_chans(df,'xltek_ch*')
                self.xltek_chans = valid_chans
                self.validation = validation
                df = df.loc[validation,:]
                neural_data_indices = self.xltek_chans['chan'].to_numpy().astype('int') - 1

                # Check if there are fewer channels than expected
                nchans = neural_data.shape[0]
                maxn = nchans - 1
                bool_idx = neural_data_indices > maxn
                if np.any(bool_idx):
                    print('-----> Expected %d channels, only %d available' % (neural_data_indices.max() + 1, nchans))
                    neural_data_indices = neural_data_indices[~bool_idx]

                neural_data = neural_data[neural_data_indices,:]
                #neural_data = neural_data[validation,:]
                self.raw_data['neural']['data'] = neural_data
            elif file_type == 'tdt':
                valid_chans, validation = self.validate_chans(df,'tdt_ch*')
                self.tdt_chans = valid_chans
                self.validation = validation
                df = df.loc[validation,:]
                neural_data_indices = self.tdt_chans['chan'].to_numpy().astype('int') - 1
                
                if neural_data.shape[0] < len(neural_data_indices):
                    
                    print(Back.YELLOW + '{:d} channels without labels removed from neural data!'.format(np.abs(np.diff([neural_data.shape[0], len(neural_data_indices)]))[0]))
                    print(Style.RESET_ALL)
                    
                    neural_data_indices = neural_data_indices[neural_data_indices 
                                                              < neural_data.shape[0]]
    
                neural_data = neural_data[neural_data_indices,:]
                self.raw_data['neural']['data'] = neural_data
            else:
                pass

            # Reset index
            df = df.reset_index(drop=True)

        # Set variables to use later
        dynamic_columns = []
        cols2rename = {}
        cols2keep = []
        dfcols = list(df.columns)
        nrows = df.shape[0]
        rowlist = list(range(nrows))

        # Go through columns
        for c in self.electrode_table['columns'].keys():
            
            col_settings = self.electrode_table['columns'][c]
            is_required = col_settings['required']

            # Find the column
            colfound = []
            if 'search' in col_settings.keys():
                r = re.compile(col_settings['search'], re.IGNORECASE)
                colfound = list(filter(r.match, dfcols))

            # IF column is found then give it the right name and fill blank cells
            if len(colfound) > 0:
                cols2rename[colfound[0]] = col_settings['title']
                if 'default' in col_settings.keys():
                    if col_settings['default'] == 999.0:
                        col_settings['default'] = np.nan
                    df[colfound[0]] = df[colfound[0]].fillna(col_settings['default'])

            # Column is required, absent and has a default
            elif (len(colfound) == 0) & is_required & ('default' in col_settings.keys()):
                df[col_settings['title']] = col_settings['default']

            # Column is required, absent
            elif len(colfound) == 0 & is_required:
                df[col_settings['title']] = 'None'
            else:
                x = 2 # Not needed

            # Make into the correct data type
            if ('type' in col_settings.keys()) & (len(colfound) != 0):
                df[colfound[0]] = df[colfound[0]].astype(col_settings['type'])

            # Append dynamic_columns and mark this column as being kept
            if (len(colfound) > 0) | is_required:
                cols2keep.append(col_settings['title'])
                dynamic_columns.append({'name': col_settings['title'], 'description': col_settings['description']})

        # Keep only the needed columns in dataframe
        df = df.rename(columns=cols2rename).loc[:,cols2keep]

        # If labels are specified then search for and use only specified labels
        if labels != None:
            df_labels = df['label'].tolist()
            isInDf = [x in labels for x in df_labels]
            df = df.loc[isInDf, :]
            df = df.reset_index(drop=True)

        self.labelfile['cleaned'] = df
        self.electrode_table['dynamic_columns'] = dynamic_columns

    # Create electrode group object
    def create_electrodeGroups(self):
        from pynwb.ecephys import ElectrodeGroup
        df = self.labelfile['cleaned']
        labels = df['label'].to_list()

        expr = '(?P<arrayID>[A-Za-z]*)(?P<arrayNum>\d+)'
        elecgroups = {}
        group_col = []
        group_name_col = []
        for elec in labels:
            elec_re = re.match(expr, elec)
            elec_array = elec_re.group('arrayID')
            if elec_array not in elecgroups.keys():
                elec_spec = df.loc[df['label'] == elec, 'spec'].iloc[0]

                if elec_spec in self.ieeg_elecs:
                    description = '%s, %s type electrodes. Recorded intracranially' % (elec_array, elec_spec)
                else:
                    description = '%s, %s type electrodes. Recorded outside the brain' % (elec_array, elec_spec)

                elecgroups[elec_array] = ElectrodeGroup(
                    elec_array,
                    description=description,
                    location='Brain',
                    device=self.device
                )
            group_col.append(elecgroups[elec_array])
            group_name_col.append(elec_array)

        df.loc[:,'group'] = group_col
        df.loc[:,'group_name'] = group_name_col
        self.labelfile['cleaned'] = df
        self.electrode_group = list(elecgroups.values())
        # return electrodeGroup

    # Convert the labelfile dataframe to HDMF5 Dynamic Table
    def labelfile2table(self):
        from pynwb.file import ElectrodeTable
        df = self.labelfile['cleaned']
        elecTable = ElectrodeTable().from_dataframe(
            df,
            self.electrode_table['name'],
            table_description=self.electrode_table['description'],
            columns=self.electrode_table['dynamic_columns']
        )
        self.electrode_table['table'] = elecTable
        return elecTable

    # Get the spec of each channel to bin
    def create_table_regions(self):
        elecTable = self.electrode_table['table']
        df = self.labelfile['cleaned']
        eegTypes = self.ieeg_elecs
        specIndices = {'ieeg': []}
        for idx, row in df.iterrows():
            if row.spec.lower() in eegTypes:
                specIndices['ieeg'].append(idx)
            else:
                if row.spec.lower() not in specIndices:
                    specIndices[row.spec.lower()] = []

                specIndices[row.spec.lower()].append(idx)

        # Create the electrode table regions
        table_regions = {'ieeg': None}
        for spec in specIndices:
            table_regions[spec] = elecTable.create_region(name='electrodes',region=specIndices[spec], description=spec + ' electrodes')

        self.electrode_table['table_regions'] = table_regions
        return table_regions

    # Create ElectricalSeries
    def create_es(self,name, data, fs, table_region):
        wrapped_data = H5DataIO(
            data=data,#data.astype('float32'),
            compression='gzip',
            compression_opts=4,
            shuffle=True
        )
        # wrapped_data = data
        from pynwb.ecephys import ElectricalSeries
        es = ElectricalSeries(
            name=name,
            data=wrapped_data,
            rate = fs,
            starting_time= 0.0,
            electrodes=table_region,
            description='Raw acquisition traces'
        )

        self.acquisitions.append(es)
        return es

    # Create TimeSeries (for other analog data such as audio or respiration)
    def create_ts(self,name,data,fs,description="None",comments="None",units='unknown'):
        wrapped_data = H5DataIO(data=data.T.astype('float32'), compression='gzip', compression_opts=4)
        # wrapped_data = data.T.astype('float32')
        from pynwb.base import TimeSeries
        ts = TimeSeries(
            name=name,
            #data = data.T,
            data=wrapped_data,
            rate=fs,
            description=description,
            comments=comments,
            starting_time=0.0,
            unit=units
        )
        self.acquisitions.append(ts)
        return ts

    # Add a note/comment at a discrete point in the recording
    def add_annotation(self,time,note):
        self.annotations['timestamps'].append(time)
        self.annotations['notes'].append(note)

    # Create LabeledEvents object
    def create_labeledevents(self):
        if len(self.annotations['timestamps']) > 0 and len(self.annotations['timestamps']) == len(self.annotations['notes']):
            from ndx_events import LabeledEvents
            events = LabeledEvents(
                name='annotations',
                description='annotations directly from recorded edf file',
                timestamps=self.annotations['timestamps'],
                labels=self.annotations['notes'],
                data=np.arange(len(self.annotations['timestamps']))
            )
            self.acquisitions.append(events)

    # Add edf file annotations to the list of annotations
    def read_edf_annotation(self,annotations):
        annotRegStr = '(?:% s)' % '|'.join(self.annots2ignore)
        for a in annotations:
            if not re.search(annotRegStr, a['description']):
                self.add_annotation(time=a['onset'], note=a['description'])

        self.create_labeledevents()

    # General function to call all the other reading neural data functions
    def read_input(self,file,eeg_chans=None,make_device=False):
        self.blockfile = os.path.basename(file).replace('.edf','')
        if self.subject != None:
            self.identifier = self.subject.subject_id 

        if file.endswith('.edf'):
            self.read_edf(file,make_device=make_device)
        elif file.endswith('.mat'):
            self.read_ecog(file,make_device=make_device)
        elif os.path.isdir(file):
            self.read_tdt(file,eeg_chans=eeg_chans,make_device=make_device)
        else:
            sys.exit('Unknown file type')

    # Read edf data
    def read_edf(self,edfFile,make_device=False):
        self.file_type = 'edf'
        edf = read_raw_edf(edfFile,preload=True)
        self.raw_data['orig'] = edf
        fs = edf.info['sfreq']
        self.raw_data['neural']['data'] = edf.get_data(units='V')
        self.raw_data['neural']['fs'] = fs
        self.raw_data['neural']['units'] = 'volts'
        self.read_edf_annotation(edf.annotations)
        edf.close()

        block = os.path.basename(edfFile)
        block,_ = os.path.splitext(block)

        settings = load_nwb_settings()
        # Get general variables
        if self.session_id == settings['meta_data']['session_id']:
            self.session_id = block

        if self.session_description == settings['meta_data']['session_description']:
            self.session_description = block

        if make_device:
            device = 'natus'
            self.create_device(device)

    # Read TDT data
    def read_tdt(self,tdtDir,eeg_chans=None,make_device=False):
        self.file_type = 'tdt'
        outList = []
        raw_data = tdt.read_block(tdtDir)
        self.raw_data['orig'] = raw_data

        # Look for the stores containing neural data. If none specified then look through a list of possibilities
        if eeg_chans == None:
            for streamList in self.tdt_eeg_chans:
                for s in streamList:
                    possibleStore = getTDTStore(raw_data, s)
                    if isinstance(possibleStore,tdt.StructType):
                        outList.append(possibleStore)

                # If there's no stores detected, skip
                if len(outList) == 0:
                    continue
                else:
                    break
        else:
            if not isinstance(eeg_chans, list): eeg_chans = list(eeg_chans)
            for store in eeg_chans:
                outList.append(getTDTStore(raw_data,store))

        # Before proceeding, make sure that all stores have the same samplerate
        fs = outList[0].fs
        samplerateMatch = np.array([x.fs == outList[0].fs for x in outList]).all()
        if not samplerateMatch:
            print('Error: EEG stores have different sampling rate!')

        # Now concatenate the data
        eegArray = outList[0].data
        if len(outList) > 1:
            for eegStore in outList[1:]:
                newStoreData = eegStore.data
                if newStoreData.shape[1] != eegArray.shape[1]:
                    minSamples = min([eegArray.shape[1], newStoreData.shape[1]])
                    eegArray = eegArray[:, 0:minSamples]
                    newStoreData = newStoreData[:, 0:minSamples]
                    
                    print(Back.YELLOW + 'EEG stores don''t have the same size! Longer one has been cut!')
                    print(Style.RESET_ALL)

                eegArray = np.concatenate((eegArray, newStoreData), axis=0)

        self.raw_data['neural']['data'] = eegArray
        self.raw_data['neural']['fs'] = fs
        self.raw_data['neural']['units'] = 'volts'

        # Get general variables
        block = os.path.basename(tdtDir)
        settings = load_nwb_settings()
        if self.session_id == settings['meta_data']['session_id']:
            self.session_id = block

        if self.session_description == settings['meta_data']['session_description']:
            self.session_description = block

        if make_device:
            device = 'tdt'
            self.create_device(device)

    # Read HBML style ecog.mat files
    def read_ecog(self,ecogFile,extra_neuraldata=None, analog_channels=None,make_device=False):
        self.file_type = 'ecog'
        df = self.labelfile['cleaned']

        # Read in the file
        ecog = read_mat(ecogFile, variable_names=['ecog'])
        ecog = ecog['ecog']
        self.raw_data['orig'] = ecog

        # Get general variables
        task = ecog['task']
        self.session_description = task
        block = os.path.basename(ecog['filename'])
        self.session_id = block

        # Set fieldtrip variable
        ftrip = ecog['ftrip']
        fs = ftrip['fsample']
        neuraldata = ftrip['trial']
        self.raw_data['neural']['labels'] = ftrip['label']
        self.raw_data['neural']['data'] = neuraldata
        self.raw_data['neural']['fs'] = fs

        if make_device:
            if ecog['filepath'].endswith('.edf'):
                device = 'natus'
            else:
                device = 'tdt'

            self.create_device(device)

        # TODO : Add in the extra neural data if specified by ecog.mat files

    # Create an ElectricalSeries for each table_region
    def regions2es(self):
        # Create an ElectricalSeries for each spec type
        table_regions = self.electrode_table['table_regions']
        data = self.raw_data['neural']['data']
        fs = self.raw_data['neural']['fs']

        for spec in table_regions.keys():
            indices = table_regions[spec].data
            indices = list(compress(indices, np.array(indices) < data.shape[0]))
            specData = data[indices, :].T.astype('float32')
            is_ieeg = spec == 'ieeg'

            # Create electrical series
            self.create_es(spec,specData, fs, table_regions[spec])

    def read_events_file(self,file):

        # Read events file based on file type
        _,ftype = os.path.splitext(file)
        if ftype == '.xlsx':
            df = pd.read_excel(file,sheet_name=0)
        elif ftype == '.csv':
            df = pd.read_csv(file)
        elif ftype == '.tsv':
            df = pd.read_csv(file,delimiter=' ')
        else:
            print('Unidentified file type')

        # Get names and descriptions of columns for DynamicTable
        nrows = df.shape[0]
        dynamicCols = []
        colnames = []

        # Search for start_time/onset and stop_time/offset columns
        for colTitle in df.columns:
            colTitle = colTitle.lower().replace(' ','_')
            if colTitle == 'onset' or colTitle == 'start_time':
                newColName = 'start_time'
                newColDesc = 'Start time of epoch, in seconds'
            elif colTitle == 'offset' or colTitle == 'stop_time':
                newColName = 'stop_time'
                newColDesc = 'Stop time of epoch, in seconds'
            else:
                colList = colTitle.split(':', 1)
                newColName = colList[0].lstrip().rstrip()
                if len(colList) == 1:
                    newColDesc = newColName
                else:
                    newColDesc = colList[1].lstrip().rstrip()


            colnames.append(newColName)
            dynamicCols.append({'name': newColName, 'description': newColDesc, 'index': list(range(nrows))})

        df.columns = colnames

        # Change all column types to object
        for label, content in df.items():
            df[label] = df[label].astype('object')

        # Change all elements to lists
        for idx, row in df.iterrows():
            for k in row.keys():
                row.at[k] = np.array([row[k]])

            df.iloc[idx, :] = row

        # Create TimeIntervals object
        from pynwb.epoch import TimeIntervals
        eventTable = TimeIntervals('trials',description='events table').from_dataframe(
            df,name='Events', table_description='Preprocessed events',columns=dynamicCols
        )
        self.events = eventTable
        return eventTable

    # Initialize NWB object
    def init_nwb(self):
        from pynwb import NWBFile
        from pynwb.epoch import TimeIntervals
        start_time = self.start_time
        if not isinstance(start_time,datetime):
            start_time = datetime.now(tzlocal())

        self.nwb = NWBFile(
            session_description = self.session_description,
            identifier = self.identifier,
            session_start_time = start_time,
            experimenter = self.experimenter,
            experiment_description = self.experiment_description,
            session_id=self.session_id,
            institution = self.institution,
            notes = self.notes,
            data_collection=self.data_collection,
            keywords = self.keywords,
            subject = self.subject,
            devices = [self.device],
            lab = self.lab,
            acquisition = self.acquisitions,
            electrodes = self.electrode_table['table'],
            electrode_groups = self.electrode_group
        )

        # If eventsTable exists, add it
        if isinstance(self.events,DynamicTable):
            event_module = self.nwb.create_processing_module(
                name='preproc_events',
                description='preprocessed events'
            )
            event_module.add(self.events)
        elif isinstance(self.events,TimeIntervals):
            self.nwb.add_time_intervals(self.events)

        if len(self.extra_files) > 0:
            files_module = self.nwb.create_processing_module(
                name='extra_files',
                description='data output directly from experiment',
                data_interfaces=self.extra_files)

    # Write out nwb data
    def write_nwb(self,output_path):
        from pynwb import NWBFile, NWBHDF5IO
        if self.nwb == None:
            self.init_nwb()

        if not output_path.endswith('.nwb'):
            output_path += '.nwb'

        if self.create_path and (not os.path.isdir(os.path.dirname(output_path))):
            os.makedirs(os.path.dirname(output_path))

        with NWBHDF5IO(output_path, mode="w") as io:
            io.write(self.nwb)

    # Parse a params dict
    #@classmethod
    def parse_params(self,params):

        if 'create_path' in params.keys():
            self.create_path = params['create_path']

        # Meta-data fields
        meta_fields = ['session_id', 'session_description', 'notes', 'data_collection', 'experimenter', 'keywords','lab','experiment_description']
        for m in meta_fields:
            if m in params.keys():
                setattr(self, m, params[m])

        # Get subject specific info
        subinfo = ['subject_id', 'sex', 'age', 'subject_description']
        subdict = {}
        for s in subinfo:
            if s in params.keys():
                subdict[s] = str(params[s])

        if subdict:
            self.create_subject(**subdict)

        # Read data
        print('-----> Reading input: %s' % params['block'])
        eeg_chans = params.get('neurodata')
        self.read_input(params['block'], eeg_chans=eeg_chans, make_device=True)

        # Update session description
        if 'experiment_description' in params.keys():
            self.session_description = params['experiment_description']
            
        if 'session_id' in params.keys():
            self.session_id = params['session_id']
        else:
            self.session_id = params['task']
        
        # Labelfile things
        if not params.get('labelfile'):
            params['labelfile'] = 'ecog'

        print('-----> Reading labelfile: %s' % params['labelfile'])
        self.read_labelfile(params['labelfile'])
        file_type = self.file_type
        labels2use = self.raw_data['neural']['labels'] if file_type == 'ecog' else None
        self.clean_labelfile(labels=labels2use)
        self.create_electrodeGroups()
        self.labelfile2table()
        self.create_table_regions()
        self.regions2es()

        # Output filename
        if params.get('output'):
            outputName = params['output']
        else:
            outputName, _ = os.path.splitext(params['block'])
            outputName = outputName + '.nwb'

        # Other acquisitions
        if params.get('analog'):
            for ana in params['analog']:
                anaName = ana['name']
                anaStore = ana.get('store') if ana.get('store') else ana.get('stores')
                if isinstance(anaStore,list): anaStore = anaStore[0]
                anaDesc = ana.get('description') if ana.get('description') else '%s acquisition' % anaName
                anaComments = ana.get('comments') if ana.get('comments') else 'None'
                chnNums = ana.get('channels') if ana.get('channels') else 999

                # Should this analog channel be written to an external file
                write2wav = False
                if 'externalize' in ana.keys():
                    tmp = str(ana['externalize']).upper()
                    write2wav = True if tmp in ['1','TRUE','YES','Y'] else False

                if write2wav:
                    print('-----> Writing to external wav file acquisition: %s' % anaName)
                else:
                    print('-----> Adding acquisition: %s' % anaName)

                # Get data depending on filetype
                if file_type == 'edf':
                    # Read raw_data and then get chans
                    chnNames = ['{:s}{:d}'.format(anaStore, ch) for ch in chnNums]
                    anaData = self.raw_data['orig'].get_data(chnNames)
                    anaFs = self.raw_data['orig'].info['sfreq']
                elif file_type == 'tdt':
                    anaTDT = getTDTStore(self.raw_data['orig'], anaStore)
                    # chnIndices = np.array(chnNums) - 1
                    chnIndices = np.arange(0, anaTDT.data.shape[0]) if chnNums == 999 else np.array(chnNums).astype(int) - 1
                    anaData = anaTDT.data[chnIndices, :]
                    anaFs = anaTDT.fs
                elif file_type == 'ecog':

                    # Parse path to the data and fs
                    prev_ext = []
                    current_ext = self.raw_data['orig']
                    for p in anaStore.split('/'):
                        prev_ext = current_ext
                        current_ext = current_ext.get(p)

                    if type(current_ext) == 'dict':
                        anaData = current_ext['trial']
                    else:
                        anaData = current_ext

                    chnIndices = np.arange(0, anaData.shape[0]) if chnNums == 999 else np.array(chnNums) - 1
                    try:
                        anaData = anaData[chnIndices, :]
                    except:
                        anaData = anaData

                    if 'fs' in prev_ext.keys():
                        anaFs = prev_ext['fs']
                    else:
                        anaFs = prev_ext['fsample']
                        
                # Write output in the case of a microphone recording
                if write2wav:

                    wav_dir, _ = os.path.split(params['mic_output'])
                    if not os.path.exists(wav_dir):
                        os.makedirs(wav_dir)
                        
                    dim1, dim2 = anaData.shape
                    if dim1 < dim2:
                        anaData = anaData.T
                         
                    # Write data to a tsv.gz file 
                    with gzip.open(params['mic_output'], 'wt') as f:    
                        for a in range(len(anaData)): 
                            f.write('{:1.20f}\t\n'.format(anaData[a][0]))             

                    # Create corresponding json file 
                    ana_descr = {
                        'SamplingFrequency': anaFs,
                        'StartTime': 0,
                        'Columns': [anaName]
                    }
                    
                    json_file = params['mic_output'].replace('_physio.tsv.gz', '_physio.json')
                    
                    with open(json_file, 'w') as file:
                        json.dump(ana_descr, file, indent=4)
    
                else:
                    self.create_ts(anaName, anaData, anaFs, description=anaDesc,comments=anaComments)

        # TDT TTL channels
        if params.get('digital'):
            for dig_chn in params['digital']:

                dig_stores = dig_chn.get('stores') if dig_chn.get('stores') else dig_chn.get('store')
                
                if file_type == 'edf':
                    
                    dig_data = self.raw_data['orig'].get_data(dig_stores)
                    event_times_store = [None] * dig_data.shape[0]
                    
                    for dc in range(dig_data.shape[0]):
                        
                        # Use Gaussian mixture models to find means 
                        gmm = GaussianMixture(n_components=3)
                        gmm.fit(dig_data[dc,:].reshape(-1, 1))
                        
                        means = gmm.means_
                        standard_deviations = gmm.covariances_**0.5  
                        
                        # Look at the fit of the first Gaussian, which is the noise
                        # Use this as a threshold
                        if 'sub-02' in outputName:
                            if 'task-rest' in outputName:
                                noise_thresh = means[0] + 6*standard_deviations[0]
                            else: 
                                noise_thresh = means[0] + 4*standard_deviations[0]
                        
                        else:
                            if 'task-rest' in outputName:
                                noise_thresh = means[0] + 4*standard_deviations[0]
                            else: 
                                noise_thresh = means[0] + 3*standard_deviations[0]
                        
                        idx_trig = dig_data[dc,:]  > noise_thresh
                        idx_onset = np.where(np.diff(idx_trig[0,:].astype(float)) > 0)[0] + 1
                        
                        s_time = self.raw_data['orig'].times[idx_onset]
                        s_name = ['{:s}'.format(dig_stores[dc]) for i in range(len(idx_onset))]
                
                        event_times_store[dc] = pd.DataFrame({'time': s_time, 'stores': s_name})
                        
                        # For debugging
                        # event_vec = np.zeros(self.raw_data['orig'].times.shape)
                        # event_vec[idx_onset] = 1
                        
                        # import matplotlib.pyplot as plt
                        # # plt.figure()
                        # plt.plot(self.raw_data['orig'].times, dig_data.T)
                        # plt.plot(self.raw_data['orig'].times, event_vec)
                        # plt.close()
                        
                    # Loop through all stores given and combine
                    event_times_df = pd.DataFrame({'time': [], 'stores': []})
                    
                    for this_store in event_times_store:
                        
                        timestamps = this_store.time
                        
                        for i,t in enumerate(timestamps):
                            idx = event_times_df['time'].isin([t])
                            if idx.any():
                                event_times_df.loc[idx, 'stores'] = event_times_df.loc[idx, 'stores'] + '/' + this_store.stores[i]
                            else:
                                event_times_df = pd.concat((event_times_df, pd.DataFrame({'time': [t], 'stores': [this_store.stores[i]]})))

                    event_times_df = event_times_df.sort_values(by='time')
                    event_times_df = event_times_df.reset_index()
                    event_times_df = event_times_df.drop('index', axis=1)
                    
                    idx = np.hstack(([False], np.diff(event_times_df.time) < (3 / self.raw_data['orig'].info['sfreq'])))
                    idx_pre = np.hstack((idx[1:], [False]))
                    
                    idx_num = np.where(idx)[0]
                    idx_pre_num = np.where(idx_pre)[0]
                    
                    store_corr = np.empty(len(idx_pre_num), dtype=object)
                    
                    for j in range(len(idx_pre_num)):
                        
                        i_idx = int(re.findall(r'-?\d+\.?\d*', event_times_df.stores.iloc[idx_num[j]])[0])
                        i_pre = int(re.findall(r'-?\d+\.?\d*', event_times_df.stores.iloc[idx_pre_num[j]])[0])
                        
                        if i_idx < i_pre:
                            store_corr[j] = '{:s}/{:s}'.format(event_times_df.stores.iloc[idx_num[j]], 
                                                               event_times_df.stores.iloc[idx_pre_num[j]])
                        elif i_idx > i_pre:
                            store_corr[j] = '{:s}/{:s}'.format(event_times_df.stores.iloc[idx_pre_num[j]], 
                                                               event_times_df.stores.iloc[idx_num[j]])
                            
                    event_times_df.loc[idx_pre, 'stores'] = store_corr
                    event_times_df = event_times_df.drop(idx_num)
                    event_times_df = event_times_df.reset_index()
                    event_times_df = event_times_df.drop('index', axis=1)
                    
                elif file_type == 'tdt':
                    event_times_df = read_tdt_ttls(self.raw_data['orig'], dig_stores)

                # Should this digital channel be written to an external file
                write2csv = False
                if 'externalize' in dig_chn.keys():
                    tmp = str(dig_chn['externalize']).upper()
                    write2csv = True if tmp in ['1','TRUE','YES','Y'] else False

                # Make sure there's actually data
                if isinstance(event_times_df,pd.DataFrame) and not write2csv:
                    print('-----> Adding acquisition: %s' % dig_chn['name'])
                    # Make numeric code
                    unique_ids = event_times_df['stores'].unique()
                    label_vals = list(range(unique_ids.size))
                    store_times = event_times_df['stores'].tolist()
                    store_codes = dict(zip(unique_ids, label_vals))
                    codes = []
                    for ii in store_times:
                        codes.append(store_codes[ii])

                    from ndx_events import TTLs
                    events = TTLs(
                        name=dig_chn['name'],
                        description=dig_chn['description'] if dig_chn.get('description') else 'TTL pulses',
                        timestamps=event_times_df['time'].to_numpy(),
                        data=codes,
                        labels=unique_ids
                    )
                    self.acquisitions.append(events)

                elif write2csv:
                    print('-----> Writing to external csv file acquisition: %s' % dig_chn['name'])
                    if outputName.endswith('_ieeg.nwb'):
                        csvFilename = outputName.replace('_ieeg.nwb','_' + dig_chn['name'] + '.csv')
                    else:
                        csvFilename = outputName.replace('.nwb','_' + dig_chn['name'] + '.csv')
                    event_times_df.to_csv(csvFilename,index=False)
                else:
                    print('-----> Error adding acquisition: %s' % dig_chn['name'])

        # Events if they exist
        if params.get('events'):
            print('-----> Reading events file: %s' % params['events'])
            self.read_events_file(params['events'])

        # Additional raw files if they exist
        if params.get('extra_files'):
            extra_files = params.get('extra_files')
            for tf in extra_files:
                tfname = os.path.basename(tf)
                print('-----> Adding %s to /processing/extra_files/%s' % (tf,tfname))
                extra_df = self.read_extra_file(tf)
                if os.path.splitext(tfname)[-1] == '.mat':
                    file_desc = 'a raw file from experiment, data are json encoded string'
                else:
                    file_desc = 'a raw file from experiment'

                self.add_file(extra_df, tfname, description=file_desc)



        print('-----> Writing file to: %s' % outputName)
        self.write_nwb(outputName)
        print('-----> Complete')

def batch_file_process(batch_excel_file,create_path=False):

    paramsdir,_ = os.path.splitext(batch_excel_file)
    paramsdir += '_params'
    if not os.path.isdir(paramsdir):
        os.mkdir(paramsdir)

    # Directory to hold all the params
    df = pd.read_excel(batch_excel_file,sheet_name='blocks',engine='openpyxl')

    # If there's a variables column then load that in
    wb = load_workbook(batch_excel_file, read_only=True, keep_links=False)
    has_var_sheet = 'variables' in wb.sheetnames
    wb.close()
    df_vars = {}
    if has_var_sheet:
        vars_df = pd.read_excel(batch_excel_file, sheet_name='variables', header=None, engine='openpyxl')
        vars_df = vars_df.astype(str)
        if not vars_df.empty:
            var_names = vars_df[0].to_list()
            var_vals = vars_df[1].to_list()
            df_vars = {var_names[ii]: var_vals[ii] for ii in range(len(var_names))}

    for idx, row in df.iterrows():

        try:

            row_dict = row.dropna().astype(str).to_dict()
            row_dict['create_path'] = create_path

            # Fill in any variables
            for var_key in df_vars.keys():
                for var_key2 in df_vars.keys():
                    var_val = df_vars[var_key2]
                    if isinstance(var_val,str):
                        var_val = var_val.replace('${' + var_key + '}', df_vars[var_key])
                        df_vars[var_key2] = var_val
            for var_key in df_vars.keys():
                for row_key in row_dict.keys():
                    row_val = row_dict[row_key]
                    if isinstance(row_val, str):
                        row_val = row_val.replace('${' + var_key + '}', df_vars[var_key])
                        row_dict[row_key] = row_val

            # replace block_id with block if necessary
            if 'block' not in row_dict.keys():
                try:
                    row_dict['block'] = df_vars['block_path'] + '/' + row_dict['block_id']
                    row_dict.pop('block_id')
                except ValueError:
                    print(
                        'blocks sheet of batch file needs to have either "block" (with the complete path to the block) or "block_id" (with the path relative to "block_path" from the variables sheet).')

            # Most important info
            blockfile = os.path.basename(row_dict['block'])

            if 'experimenter' in row_dict.keys(): row_dict['experimenter'] = row_dict['experimenter'].split(';')

            # Neurodata
            if 'neurodata' in row_dict.keys():
                tmp = row_dict['neurodata'].split(',')
                row_dict['neurodata'] = [x.rstrip().lstrip() for x in tmp]
                #row_dict['neurodata'] = row_dict['neurodata'].split(',')

            # Analog channels
            analog_prefixes = ['analog1','analog2','analog3','analog4','analog5','analog6']
            fields2add = ['name', 'store', 'channels', 'description', 'comments', 'externalize']
            analist = []
            for a in analog_prefixes:
                # If name doesn't exist, then it isn't there
                ana_name = a + '_name'
                if ana_name not in row_dict.keys():
                    continue
                else:
                    new_ana = {}
                    for f in fields2add:
                        field2find = a + '_' + f
                        if field2find in row_dict.keys():
                            if 'externalize' in field2find:
                                new_ana[f] = bool(int(float(row_dict[field2find])))
                            else:
                                new_ana[f] = row_dict[field2find]
                            row_dict.pop(field2find)

                    if 'channels' in new_ana.keys():
                        new_ana['channels'] = [int(float(x)) for x in new_ana.get('channels').split(',')]

                    analist.append(new_ana)

            # Digital channels
            digital_prefixes = ['digital1','digital2','digital3']
            fields2add = ['name', 'stores', 'description', 'comments']
            diglist = []
            for d in digital_prefixes:
                dig_name = d + '_name'
                if dig_name not in row_dict.keys():
                    pass
                else:
                    new_dig = {}
                    for f in fields2add:
                        field2find = d + '_' + f
                        if field2find in row_dict.keys():
                            new_dig[f] = row_dict[field2find]
                            row_dict.pop(field2find)

                    new_dig['stores'] = new_dig['stores'].split(',')
                    diglist.append(new_dig)

            # Define output Path
            if 'output' not in row_dict.keys():
                if 'task' in row_dict.keys():
                    outputName = '{:s}_{:s}_task-{:s}'.format(df_vars['subject_id_bids'],
                                                              df_vars['session_id'],
                                                              row_dict['task'])
                    if 'acq' in row_dict.keys():
                        outputName = outputName + '_acq-' + row_dict['acq']
                    if 'run' in row_dict.keys():
                        outputName = '{:s}_run-{:02d}'.format(outputName, 
                                                              int(row_dict['run']))
                else:
                    outputName, _ = os.path.splitext(blockfile)
                    outputName = outputName + '.nwb'
                if 'output_path' in df_vars.keys():
                    outputName = df_vars['output_path'] + '/' + df_vars['session_id'] + '/ieeg/' + outputName
                if not outputName.endswith('_ieeg.nwb'):
                    outputName += '_ieeg.nwb'
                row_dict['output'] = outputName
                
            # Create output for external wave file (microphone)
            idx_mic = np.where([ana['name'] == 'mic' for ana in analist])[0]
            
            if len(idx_mic) == 1:
                
                if row_dict['output'].endswith('_ieeg.nwb'):
                    row_dict['mic_output'] = row_dict['output'].replace('_ieeg.nwb', '_physio.tsv.gz')
                
            # add subject level data if necessary
            if 'subject_id' not in row_dict.keys():
                if 'subject_id_bids' in df_vars.keys():
                    row_dict['subject_id'] = df_vars['subject_id_bids']
                else:
                    print('subject ID could not be found. Either the main sheet or the variables sheet of the batch file need to have a field called "subject_id"')
            if 'sex' not in row_dict.keys():
                if 'sex' in df_vars.keys():
                    row_dict['sex'] = df_vars['sex']
                else:
                    print('subject ID could not be found. Either the main sheet or the variables sheet of the batch file need to have a field called "sex"')
            if 'age' not in row_dict.keys():
                if 'age' in df_vars.keys():
                    row_dict['age'] = df_vars['age']
                else:
                    print('subject ID could not be found. Either the main sheet or the variables sheet of the batch file need to have a field called "age"')
            if 'subject_description' not in row_dict.keys():
                if 'subject_description' in df_vars.keys():
                    row_dict['subject_id'] = df_vars['subject_id_bids']
                else:
                    print('subject description could not be found. Either the main sheet or the variables sheet of the batch file should have a field called "subject_description"')
            if 'labelfile' not in row_dict.keys():
                if 'corr_sheet' in df_vars.keys():
                    row_dict['labelfile'] = df_vars['corr_sheet']
                else:
                    print('electrode correspondence sheet could not be found. Either the main sheet or the variables sheet of the batch file should have a field called "corr_sheet"')

            # Add digital analog
            subfields = zip(['analog', 'digital'], [analist, diglist])
            for k,v in subfields:
                if len(v) > 0:
                    row_dict[k] = v

            # Create a yml file
            outfile = paramsdir + os.sep + '%s.yml' % blockfile.replace('.edf', '')
            with open(outfile, 'w') as file:
                yaml.dump(row_dict, file, sort_keys=False)

            print('*' * 150)
            print('Processing %s' % blockfile)
            print('Params written to %s' % outfile)
            print('*' * 150)

            # Parse
            inwb = ieeg2nwb()
            inwb.parse_params(row_dict)

        except:
            print('*' * 200)
            print('Error processing %s. Skipping to next file' % blockfile)
            print('*' * 200)

def cmnd_line_parser():
    # Create parser
    parser = argparse.ArgumentParser(description="Convert a file to NWB format",
                                    epilog=example_usage + additional_notes,
                                    formatter_class=argparse.RawDescriptionHelpFormatter
                                    )
    parser.add_argument('--batch', required=False, help='excel file for batch conversion',dest='batch_file',default=None)
    parser.add_argument('--gui', required=False, help='launch the ieeg2nwb gui',dest='gui',action='store_true')
    parser.add_argument('--params','-p', required=False, help='json or yml params file to use instead of command line arguments',dest='params_file',default=None)
    parser.add_argument("--block_path", "-b", required=False, help="Path to the data to convert to NWB", dest='block')
    parser.add_argument("--labelfile", "-l", required=False, help="Path to the correspondence sheet", dest='labelfile')
    parser.add_argument("--output", "-o", required=False, default=None, dest='output',
                        help="Where to write the output .nwb file to. Defaults to where block path was read"
                        )
    parser.add_argument("--events", "-e", dest='events', default=None,
                        help="""
                        A csv, tsv or xlsx file that contains the events (if any) for this recording
                        Columns must include description of its' contents using a colon.
                        Ex:
                        Trial : The trial of this point, Onset : The time of stimulus onset
                        """
                        )
    parser.add_argument("--analog", "-a", nargs="+", dest='analog',
                        help="""Any analog channels recorded alongside the experiment. Must include channel located and name of it
                        Ex:
                        --analog Mic:Microphone recorded during task:Wav5:1,2 <-- Creates a timeseries called "Mic" that is located in Wav5 channels 1 and 2
                        --analog Sound:Audio Recorded:DC1 Mic:Microphone from task:DC2 <-- Creates two timeseries containers with data from channels DC1 and DC2
                        """
                        )
    parser.add_argument("--digital", "-d", nargs="+", dest='digital', default=None,
                        help="""
                        Creates container that stores discrete timepoints. Used for TTL pulses. If TTL pulses were recorded in an analog channel then please use the --analog option
                        Ex:
                        --digital TTL:PtC4 <-- Creates a container to store discrete timepoints in the PtC4 store
                        --digital TTL1:PtC2 TTL2:PtC4,PtC6 <-- Creates two separate containers for TTL pulses. TTL2 contains the discrete timepoints from both the PtC4 and PtC6 stores
                        """
                        )
    parser.add_argument("--neurodata", "-n", nargs="+", dest='eeg_chans', default=None,
                        help="""
                        Used to specify where the neural data is stored when input is a TDT block. If unspecified and input is TDT then the following stores will be attempted to retrieve and extract from (in order):
                        - RAWx
                        - EEG1 and EEG2
                        - RSn1 and RSn2
                        """
                        )
    parser.add_argument('--session_id', help="Name of the session. By default will be the block filename. Must be surrounded in quotes", dest='session_id', default=None)
    parser.add_argument('--age', help='Age of the subject', dest='age', default=None)
    parser.add_argument('--subid', '--subject_id', help='Name of the subject', dest='subject_id', default=None)
    parser.add_argument('--sex', help='Sex of the participant', dest='sex', default=None)
    parser.add_argument('--subject_description', help='short description of the subject, must be surrounded in quotes', dest='subject_description', default=None)
    parser.add_argument('--session_description', help='description of session, must be surrounded in quotes', dest='session_description', default=None)
    parser.add_argument('--experiment_description', help='description of experiment, must be surrounded in quotes', dest='experiment_description', default=None)
    parser.add_argument('--notes', help='any additional notes, must be surrounded in quotes', dest='notes', default=None)
    parser.add_argument('--data_collection', help='any additional notes about data collection, must be surrounded in quotes', dest='data_collection', default=None)
    parser.add_argument('--experimenters', nargs="+", help='Any people involved in administering the task. Must be surrounded by quotes', dest='experimenters', default=None)
    parser.add_argument('--keywords', nargs="+", help='Any additional keywords to add to this file', dest='keywords', default=None)
    parser.add_argument('--text-files', nargs="+", help='Raw text files to store in the nwb file', dest='text_files',default=None)
    parser.add_argument('--lab', nargs="+", help='Lab(s) the data should be attributed to', dest='lab', default=None)
    parser.add_argument('--mkdirs', required=False, help='create the full output directory path if it does not exist',dest='create_path',action='store_true')
    args = parser.parse_args()

    # Setup params
    params = vars(args)

    # Check if params or block path is passed in
    if params['params_file'] == None and params['block'] == None and params['batch_file'] == None and params['gui'] == False:
        print('Error! Have to specify params file OR block with recorded dat')
        parser.print_help()
        sys.exit(2)


    if params['gui']:
        from PyQt5.QtWidgets import QApplication
        from .gui import GUI
        app = QApplication([])
        ex = GUI()
        ex.show()
        sys.exit(app.exec_())


    if params['params_file'] != None and os.path.isfile(params['params_file']):
        with open(params['params_file']) as file:
            params = yaml.load(file, Loader=yaml.FullLoader)

        inwb = ieeg2nwb()
        inwb.parse_params(params)

    elif params['batch_file'] != None and os.path.isfile(params['batch_file']):
        batch_file_process(params['batch_file'],create_path=params['create_path'])

    else:
        # Check for additional channels
        if params.get('analog'):
            analogChns = params['analog']
            params['analog'] = []
            for storageString in analogChns:
                anaList = storageString.split(':')
                new_ana = {
                    'name': anaList[0],
                    'description': anaList[1],
                    'store': anaList[2]
                }
                if len(anaList) == 4:
                    new_ana['channels'] = [int(x) for x in anaList[3].split(',')]

                params['analog'].append(new_ana)


        # Digital channels (TTLs in TDT)
        if params.get('digital'):
            digitalContainers = params['digital']
            params['digital'] = []
            for container in digitalContainers:
                container_items = container.split(':')
                if len(container_items) == 3:
                    containerName, containerDescription, containerSources = container_items
                else:
                    containerName, containerSources = container_items
                    containerDescription = 'TTL pulses'

                containerSources = containerSources.split(',')
                params['digital'].append({'name': containerName,'stores': containerSources,'description': containerDescription})

        inwb = ieeg2nwb()
        inwb.parse_params(params)


# If this is run at command line then run this whole thing
if __name__ == "__main__":
    cmnd_line_parser()
